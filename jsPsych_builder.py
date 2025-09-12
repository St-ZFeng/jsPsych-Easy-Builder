import reflex as rx
import json,random,string
import copy
from typing import Any
import os
from .jsPsych_trans import Expriment, registry_plugin,remove_dollar
import pathlib
import zipfile
from pathlib import Path
import shutil
import csv
from fastapi import FastAPI, Request
from openpyxl import load_workbook



mode='Local'

if mode == 'Local':
    data_upload = FastAPI()

    @data_upload.post("/receive-csv/")
    async def receive_csv(request: Request):
        raw_csv_data = await request.body()
        csv_text = raw_csv_data.decode('utf-8')

        if 'trial_index' in csv_text:
            data_dir = "data"
            os.makedirs(data_dir, exist_ok=True) 

            def generate_unique_csv_filename(directory):
                while True:
                    random_str = ''.join(random.choices(string.ascii_lowercase + string.digits, k=8))
                    filename = f"upload_{random_str}.csv"
                    file_path = os.path.join(directory, filename)
                    if not os.path.exists(file_path):
                        return file_path

            save_path = generate_unique_csv_filename(data_dir)
            with open(save_path, "w", encoding="utf-8") as f:
                f.write(csv_text)
 



circle_style = {
    "width": "24px",
    "height": "24px",
    "borderTop": "4px solid #3498db",          
    "borderRight": "4px solid #3498db",       
    "borderBottom": "4px solid #3498db",      
    "borderLeft": "4px solid transparent",    
    "borderRadius": "50%",
    "animation": "spin 1s linear infinite",
    "display": "inline-block",
}

css_keyframes = """
@keyframes spin {
    0% { transform: rotate(0deg);}
    100% { transform: rotate(360deg);}
}
"""
plugin_registry=registry_plugin()

plugin_dict={}
params_dict={}
simple_params={}
exp_test = Expriment("Expriment_test")
version = exp_test.version
def trans_to_form_param(type,value):
    if type=='bool' and value==None:
        return False
    elif type=='bool' and isinstance(value, bool):
        return value
    elif type=='bool' and value in ("true","false"):
        return bool(value.title())
    elif value==None:
        return ""
    else:
        return str(value)
def trans_to_form_date(plugin_dict,simple):
    plugin_name=simple['type']
    for param in simple:
        if param in ('name','type'):
            continue
        simple[param]=trans_to_form_param(plugin_dict[plugin_name].params[param]["type"],simple[param])
    return simple


def file_to_json_list(file_path):
    file_ext = os.path.splitext(file_path)[1].lower()
    
    if file_ext == '.csv':
        return read_csv(file_path)
    elif file_ext == '.xlsx':
        return read_excel(file_path)
    else:
        raise ValueError("Only for .csv and .xlsx")

def read_csv(file_path):
    with open(file_path, 'r', encoding='utf-8-sig') as csv_file:
        reader = csv.DictReader(csv_file)
        return [row for row in reader]

def read_excel(file_path):
    wb = load_workbook(filename=file_path, read_only=True)
    sheet = wb.active
    
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    
    json_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, value in enumerate(row):
            if value is None:
                value = ""
            row_dict[headers[i]] = str(value) if value is not None else ""
        json_list.append(row_dict)
    
    wb.close()
    return json_list

    
for plugin_name in plugin_registry:
    plugin_dict[plugin_name]=plugin_registry[plugin_name](exp_test)
    params_dict[plugin_name]={"type":plugin_name,"name":plugin_name.replace("-","_"),"params":plugin_dict[plugin_name].params}
    simple={}
    #print('---------------------------------------------')
    #print(plugin_dict[plugin_name].params)
    for param in plugin_dict[plugin_name].params:
        
        simple[param]=plugin_dict[plugin_name].params[param]["value"]
    simple["name"]=plugin_name.replace("-","_")
    simple["type"]=plugin_name
    simple_params[plugin_name]=trans_to_form_date(plugin_dict,simple)
    simple_params[plugin_name]["disabled"]=False

class Preview_state(rx.State):
    iframe_key: int = 0 
    def refresh_iframe_simple(self):
        self.iframe_key=self.iframe_key+1

class Csv_state(rx.State):
    current_csv:str=''


    async def clear(self):
        if self.current_csv!='':
            user_floder=await self.get_var_value(State.user_floder)
            outfile = user_floder / self.current_csv
            outfile.unlink()
            self.current_csv=''

    async def handle_csv_upload(
        self, files: list[rx.UploadFile]
    ):
        self.current_csv=files[0].name
        for file in files:
            upload_data = await file.read()

            user_floder=await self.get_var_value(State.user_floder)

            outfile = user_floder / file.name

            with outfile.open("wb") as file_object:
                file_object.write(upload_data)


class Uploaded_files_state(rx.State): 
    files: list[str]

    async def check_uploaded_file(self):
        user_floder=await self.get_var_value(State.user_floder)
        files = list(user_floder.iterdir())
        files = [item for item in files if item.is_file()]
    
        filtered_files = [
            file for file in files
            if file.name not in ['dist','index.html','index_preview.html','setting.jsexp']
        ]
        filtered_files = [str(file.name) for file in filtered_files if file.suffix.lower() != '.zip']
        self.files=filtered_files
    

    async def delete_file(self,name,id):
        if 0 <= id < len(self.files) and name in self.files:
            user_floder=await self.get_var_value(State.user_floder)
            del self.files[id]
            file_path=user_floder / name
            if file_path.exists():
                file_path.unlink()
    async def delete_all_file(self):
        user_floder=await self.get_var_value(State.user_floder)
        for file in self.files:
            file_path=user_floder/ file
            if file_path.exists():
                file_path.unlink()
        self.files=[]

    

    @rx.event
    async def handle_upload(
        self, files: list[rx.UploadFile]
    ):
        user_floder=await self.get_var_value(State.user_floder)
        for file in files:
            upload_data = await file.read()
            outfile = user_floder / file.name

            # Save the file.
            with outfile.open("wb") as file_object:
                file_object.write(upload_data)

            if file.name not in self.files:
                self.files.append(file.name)
    
class complex_temp_state(rx.State):
    complex_list_temp:list[dict[str, Any]]=[]

    async def add_complex_temp(self,plugin_name,param_name,index=None):
        complex_source=params_dict[plugin_name]['params'][param_name]['nested']
        complex_temp={}
        for param in complex_source:
            if param=='required':
                complex_temp[param]=True
            else:
                complex_temp[param]=trans_to_form_param(complex_source[param]['type'],complex_source[param]['default'])
        if index==None:
            self.complex_list_temp.append(complex_temp)
        else:
            complex_list=await self.get_var_value(State.complex_list)
            complex_list[index].append(complex_temp)
    
    async def delete_complex_temp(self, complex_id,id):
        if id==None:
            del self.complex_list_temp[complex_id]
        else:
            complex_list=await self.get_var_value(State.complex_list)
            del complex_list[id][complex_id]

    def clear_complex_temp(self):
        if len(self.complex_list_temp)!=0:
            self.complex_list_temp=[]

class State(rx.State):
    user_floder:pathlib.Path= rx.get_upload_dir()/'jspsych_exp'
    first_open:bool = True
    simple_params_state:dict[str,dict[str, Any]] = copy.deepcopy(simple_params)
    preview_start: str = 'index_0'
    timeline: list[dict[str, Any]] = []
    complex_list: list[list[dict[str, Any]]] = []
    building_exp_file:bool=False
    exp_file:str =''
    false:str =False
    true:str = True
    name_used:list[str] =['jsPsych','timeline']


    exp_name: str ='Experiment'
    plugin_source: str ='Local'
    data_save: str='Local'
    head_script: str=''
    procedure_list:list[str]
    Unterminated_procedure_list:list[str]
    preview_file:str ='jspsych_exp/index_preview.html'
    preview_html_default:str ="""
        <!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8" />
  <title>Preview</title>
  <style>
    html, body {
      height: 100%;
      margin: 0;
    }
    body {
      display: flex;
      justify-content: center; 
      align-items: center;    
      text-align: center;
      height: 100vh;
    }
  </style>
</head>
<body>
  <h1>Add plugins to preview here.</h1>
</body>
</html>
    """
    
    def add_complex(self,index,complex_temp):
        self.complex_list[index].append(complex_temp)

    def delete_complex(self,complex_id,id):
        del self.complex_list[id][complex_id]

    @rx.event
    async def set_user_floder(self):
        user_dir = rx.get_upload_dir()
        user_dir.mkdir(parents=True, exist_ok=True)  
        if self.first_open:
            self.first_open=False
            if mode=="Local":
                folder_name ='jspsych_exp'
                random_folder = user_dir / folder_name
                if not random_folder.exists():
                    random_folder.mkdir()
                self.user_floder=random_folder
            else:
                while True:
                    folder_name = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
                    random_folder = user_dir / folder_name
                    if not random_folder.exists():
                        random_folder.mkdir()
                        self.user_floder=random_folder
                        break
            src = pathlib.Path("./dist")
            dst = self.user_floder/ "dist"
            dst.mkdir(parents=True, exist_ok=True)
            for item in src.iterdir():
                if item.is_dir():
                    shutil.copytree(item, dst / item.name, dirs_exist_ok=True)
                else:
                    shutil.copy2(item, dst / item.name)
            setting_file=random_folder / 'setting.jsexp'
        
            if setting_file.exists():
                await self.load_exp_json()
            await self.refresh_iframe()
            self.preview_file= folder_name+'/index_preview.html'
            
        else:
            if not self.user_floder.exists():
                self.user_floder.mkdir()
                src = pathlib.Path("./dist")
                dst = self.user_floder/ "dist"
                
                dst.mkdir(parents=True, exist_ok=True)
                for item in src.iterdir():
                    if item.is_dir():
                        shutil.copytree(item, dst / item.name, dirs_exist_ok=True)
                    else:
                        shutil.copy2(item, dst / item.name)
                await self.refresh_iframe()
                self.preview_file= folder_name+'/index_preview.html'

    @rx.event
    async def new_exp(self):
        self.timeline=[]
        self.complex_list=[]
        self.name_used =['jsPsych','timeline']  
        self.head_script=''
        self.procedure_list=[]
        self.Unterminated_procedure_list=[]
        await self.refresh_iframe()

    def update_next_name(self,plugin_type):
        start=0
        while True:
            if start ==0:
                next_name=plugin_type.replace('-',"_")
            else:
                next_name=plugin_type.replace('-',"_")+f'_{start}'
            start=start+1
            if next_name not in self.name_used:
                self.simple_params_state[plugin_type]["name"]=next_name
                break

    def timeline_push(self,point,complex_i,id=None):
        plugin_name=point['type']
        if plugin_name not in self.simple_params_state.keys():
            return
        if id==None:
            self.timeline.append(point)
            self.complex_list.append(complex_i)
        else:
            self.timeline.insert(id,point)
            self.complex_list.insert(id,complex_i)

        
        if plugin_name=='procedure-start':
            self.procedure_list.append(point["name"])
            self.Unterminated_procedure_list.append(point["name"])
        elif plugin_name=='procedure-end':
            self.Unterminated_procedure_list=[x for x in self.Unterminated_procedure_list if x != point['connect']] 
        self.name_used.append(point["name"])
        
    def timeline_delete(self,id,for_edit=False):
        if self.timeline[id]['type'] == 'procedure-start':
            if not for_edit:
                for i in range(len(self.timeline)):
                    if self.timeline[i]['type'] == 'procedure-end' and self.timeline[i]['connect']==self.timeline[id]['name']:
                            self.timeline_delete(i)
                   
            self.procedure_list=[x for x in self.procedure_list if x != self.timeline[id]['name']]
            self.Unterminated_procedure_list=[x for x in self.Unterminated_procedure_list if x != self.timeline[id]['name']]
        elif self.timeline[id]['type'] == 'procedure-end':
            self.Unterminated_procedure_list.append(self.timeline[id]['connect'])
        self.name_used=[x for x in self.name_used if x != self.timeline[id]['name']]
        del self.timeline[id]
        del self.complex_list[id]
    
    def timeline_edit(self,point,complex_i,id):
        name=self.timeline[id]['name']
        self.timeline_delete(id,True)
        self.timeline_push(point,complex_i,id)
        plugin_name=point['type']
        if plugin_name=='procedure-start':
            for i in range(len(self.timeline)):
                if self.timeline[i]['type'] == 'procedure-end' and self.timeline[i]['connect']==name:
                    self.timeline[i]['connect']=point["name"]
                    self.timeline_edit(self.timeline[i],[],i)





    @rx.event
    async def load_exp_json(self, file_path=None):
        

        if file_path is None:
            filename = "setting.jsexp"
            file_path = self.user_floder / filename
        if file_path.exists():
            json_data = json.loads(file_path.read_text(encoding='utf-8'))
            self.exp_name = json_data.get('exp_name')
            self.plugin_source = json_data.get('plugin_source')
            self.data_save = json_data.get('data_save')
            self.head_script = json_data.get('head_script')

            self.timeline=[]
            self.complex_list=[]
            self.name_used =['jsPsych','timeline']  
            self.procedure_list=[]
            self.Unterminated_procedure_list=[]

            self.simple_params_state= copy.deepcopy(simple_params)
            for point,complex_i in zip(json_data.get('timeline'),json_data.get('complex_list')):
                self.timeline_push(point,complex_i)
                self.update_next_name(point['type'])

        uploaded_files_state=await self.get_state(Uploaded_files_state)
        await uploaded_files_state.check_uploaded_file()
        await self.refresh_iframe()



    def save_exp_json(self):
        exp_json={}
        exp_json['exp_name']=self.exp_name
        exp_json['plugin_source']=self.plugin_source
        exp_json['data_save']=self.data_save
        exp_json['head_script']=self.head_script
        exp_json['timeline']=self.timeline
        exp_json['complex_list']=self.complex_list
        filename = "setting.jsexp"
        file_path = self.user_floder / filename
        file_path.write_text(json.dumps(exp_json, ensure_ascii=False, indent=4), encoding='utf-8')

       

    @rx.event
    async def save_exp_settings(self,setting_json):
        # Save values into the main State
        self.exp_name = setting_json['exp_name']
        self.plugin_source =setting_json['plugin_source']
        self.head_script = setting_json['head_script']
        self.data_save = setting_json['data_save']
        await self.refresh_iframe()

    def set_preview_start(self, value):
        if value=='' or value==None:
            value='index_0'
        self.preview_start = value

    def create_preview_file(self,preview_html):
        filename = "index_preview.html"
        file_path = self.user_floder / filename
        file_path.write_text(preview_html, encoding="utf-8")

    def zip_folder(self, folder_path, output_zip, excluded_items=None, excluded_extensions=None):
        if excluded_items is None:
            excluded_items = []
        if excluded_extensions is None:
            excluded_extensions = []
        
        with zipfile.ZipFile(output_zip, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(folder_path):
                dirs[:] = [d for d in dirs if str(Path(root).relative_to(folder_path) / d) not in excluded_items]
                
                for file in files:
                    file_path = Path(root) / file
                    relative_path = file_path.relative_to(folder_path)
                    relative_str = str(relative_path)
                    
                    should_exclude = False
                    
                    for excluded_item in excluded_items:
                        if relative_str == excluded_item or relative_str.startswith(excluded_item + '/'):
                            should_exclude = True
                            break
                    
                    if file_path.suffix.lower() in [ext.lower() for ext in excluded_extensions]:
                        should_exclude = True
                    
                    if not should_exclude:
                        zipf.write(file_path, relative_path)

    def save_exp_file(self):
        if self.building_exp_file==False:
            self.building_exp_file=True
            zip_files = list(self.user_floder.glob("*.zip"))
            for zip_file in zip_files:
                os.remove(zip_file)
            filename = "index.html"
            timeline_date=[]
            for timepoint in self.timeline:
                timeline_date.append(self.timeline_temp_to_date(timepoint))
            my_exp=Expriment(self.exp_name,plugin_all=plugin_registry, timeline_list=timeline_date,data_save=self.data_save,plugin_source=self.plugin_source,head_script=self.head_script)
            html_text=my_exp.export()
            file_path = self.user_floder / filename
            file_path.write_text(html_text, encoding="utf-8")
            exp_file=str(self.user_floder)+f'/{self.exp_name}.zip'
            parts = self.user_floder.parts
            self.exp_file= str(Path(*parts[1:]))+f'/{self.exp_name}.zip'

            if self.plugin_source!='Local':
                self.zip_folder(self.user_floder ,exp_file,['index_preview.html',f'{self.exp_name}.zip','dist'])
            else:
                self.zip_folder(self.user_floder ,exp_file,['index_preview.html',f'{self.exp_name}.zip'])
            self.building_exp_file=False
        
        


    @rx.event
    async def refresh_iframe(self):
        if len(self.timeline)!=0:
            
            preview_start=int(self.preview_start.replace('index_',''))
            if preview_start<0 :
                preview_start=0
                self.preview_start='index_0'
            elif preview_start>=len(self.timeline):
                preview_start=len(self.timeline)-1
                self.preview_start=f'index_{preview_start}'
            timeline_date=[]
            
            for timepoint in self.timeline:
                timeline_date.append(self.timeline_temp_to_date(timepoint))
            #print(timeline_date)
            my_exp=Expriment(self.exp_name,plugin_all=plugin_registry, timeline_list=timeline_date,data_save=self.data_save,plugin_source=self.plugin_source,head_script=self.head_script)
            self.create_preview_file(my_exp.preview(preview_start))
            
        else:
            self.create_preview_file(self.preview_html_default)
        await self.refresh_iframe_simple()
        self.save_exp_json()

    @rx.event
    async def refresh_iframe_simple(self):
        preview_state=await self.get_state(Preview_state)
        preview_state.refresh_iframe_simple()

    



    def timeline_param_to_date(self,param_value,param_type,param_default,param_array):
        if param_type=="bool":
            if type(param_value) ==str and "$" not in param_value and '[' in param_value:
                param_value=eval(param_value)
        elif param_type=="free":
            if param_value=='False':
                param_value=False
            elif param_value=='True':
                param_value=True
            elif param_value=='false':
                param_value=False
            elif param_value=='true':
                param_value=True
            else:
                param_value=eval(param_value)
        elif param_value=="" and param_default==None:
            param_value=None
        elif param_type=="int" and "[" not in param_value and "$" not in param_value:
            if param_value.lstrip("+-").isdigit():
                param_value=int(param_value)
            else:
                param_value=param_default
        elif param_type=="float" and "[" not in param_value and "$" not in param_value:
            def is_float(s):
                try:
                    float(s)
                    return True
                except ValueError:
                    return False
            if is_float(param_value):
                param_value=float(param_value)
            else:
                param_value=param_default
            
        elif "[" in param_value and "$" not in param_value and param_type not in ("object","function"):
            param_value=eval(param_value)
        elif "[" in param_value and "$" not in param_value and param_type in ("object","function") and param_array:
            param_value=param_value
        return param_value
    

    def timeline_temp_to_date(self,param_temp):
        param_temp=copy.deepcopy(param_temp)
        plugin_name=param_temp['type']
        param_source=params_dict[plugin_name]
        for param in param_temp:
            if param in ('name','type'):
                continue
            elif param == 'disabled':
                param_temp[param]=param_temp[param]
                continue
            elif param_source["params"][param]['type']!='complex':
                param_temp[param]=self.timeline_param_to_date(param_temp[param],param_source["params"][param]['type'],param_source["params"][param]['default'],param_source["params"][param]['array'])
            else:
                for i in range(len(param_temp[param])):
                    for sub_param in param_temp[param][i]:
                        param_temp[param][i][sub_param]=self.timeline_param_to_date(param_temp[param][i][sub_param],param_source["params"][param]['nested'][sub_param]['type'],param_source["params"][param]['nested'][sub_param]['default'],param_source["params"][param]['nested'][sub_param]['array'])

 
        return param_temp
    
    

    @rx.event
    async def add_plugin(self,plugin_name, form_data: dict):
        param_source=params_dict[plugin_name]
        param_temp=copy.deepcopy(simple_params[plugin_name])
        sorted_complex=[]
        if form_data["name"]!="":
            param_temp["name"]=form_data["name"]
        
        for param in param_temp:
            if param in ('name','type'):
                continue
            if param == 'disabled':
                continue
            if param_source["params"][param]['type']!='complex':
                if param_source["params"][param]['type']=="bool":

                    if param not in form_data:
                        param_temp[param]=False
                    elif form_data[param]=='False':
                        param_temp[param]=False
                    elif form_data[param]=='True':
                        param_temp[param]=True
                    elif form_data[param]=='on':
                        param_temp[param]=True
                    else:
                        param_temp[param]=str(form_data[param])
                elif (param=='timeline_variables' or param=='data_value') and (form_data[param].endswith('.csv') or form_data[param].endswith('.xlsx')):
                    csv_file = self.user_floder / form_data[param]
                    var_list=file_to_json_list(str(csv_file))
                    changed_var_list=[]
                    for var in var_list:
                        dict_temp={}
                        for column in var:
                            dict_temp['$'+column]=var[column]
                        changed_var_list.append(dict_temp)
                    val_str=remove_dollar(str(changed_var_list))
                    param_temp[param]=val_str
                    current_csv=await self.get_var_value(Csv_state.current_csv) 
                    current_csv=''

                
                else:
                    param_temp[param]=str(form_data[param])
            else:
                result = {}
                for key, value in form_data.items():
                    if key.startswith(param + '_'):
                        remaining_part = key[len(param)+1:]
                        parts = remaining_part.split('_')
                        
                        if len(parts) >= 2:
                            try:
                                num = str(parts[-1])
                            except ValueError:
                                continue 
                            new_key = '_'.join(parts[:-1])
                            
                            if num not in result:
                                result[num] = {}

                            if param_source["params"][param]['nested'][new_key]['type']=="bool":
                                if value=='False':
                                    value=False
                                elif value=='True':
                                    value=True
                                elif value=='on':
                                    value=True
                                else:
                                    value=str(value) 
                            else:
                                value=str(value)

                            result[num][new_key] = value
                
                for id in result:
                    for sub_param in param_source["params"][param]['nested']:
                        if sub_param not in result[id]:
                            result[id][sub_param] = False
                sorted_complex = [result[id] for id in sorted(result.keys())]
                param_temp[param]=sorted_complex
        
        point=int(form_data['add_or_move_point'].replace('index_',''))
        if point==-1:#'At the end':
            self.timeline_push(param_temp, sorted_complex)
        else:
            self.timeline_push(param_temp, sorted_complex,point)
            #before_name=form_data['add_or_move_point'].replace('Before the ',"")
            #for i in range(len(self.timeline)):
                #if self.timeline[i]['name']==before_name:
                    #self.timeline_push(param_temp, sorted_complex,i)
                    #break
        
        self.update_next_name(param_temp['type'])
        

        await self.refresh_iframe()

    @rx.event
    async def delete_plugin(self, id):
        if 0 <= id < len(self.timeline):
            plugin_type=self.timeline[id]['type']
            self.timeline_delete(id)
            self.update_next_name(plugin_type)
        await self.refresh_iframe()

    @rx.event
    async def edit_plugin(self,plugin_name, form_data: dict,index:int):
        param_source=params_dict[plugin_name]
        param_temp=copy.deepcopy(simple_params[plugin_name])
        sorted_complex=[]
        if form_data["name"]!="":
            param_temp["name"]=form_data["name"]
        else:
            pass
        
        for param in param_temp:
            if param in ('name','type'):
                continue
            if param == 'disabled':
                if param not in form_data:
                    param_temp[param]=False
                elif form_data[param]=='on':
                    param_temp[param]=True
                else:
                    param_temp[param]=False
                continue
            if param_source["params"][param]['type']!='complex':
                if param_source["params"][param]['type']=="bool":
                    if param not in form_data:
                        param_temp[param]=False
                    elif form_data[param]=='False':
                        param_temp[param]=False
                    elif form_data[param]=='True':
                        param_temp[param]=True
                    elif form_data[param]=='on':
                        param_temp[param]=True
                    else:
                        param_temp[param]=str(form_data[param])
                elif (param=='timeline_variables' or param=='data_value') and (form_data[param].endswith('.csv') or form_data[param].endswith('.xlsx')):
                    csv_file = self.user_floder / form_data[param]
                    var_list=file_to_json_list(str(csv_file))
                    changed_var_list=[]
                    for var in var_list:
                        dict_temp={}
                        for column in var:
                            dict_temp['$'+column]=var[column]
                        changed_var_list.append(dict_temp)
                    val_str=remove_dollar(str(changed_var_list))
                    param_temp[param]=val_str
                    csv_state=await self.get_state(Csv_state) 
                    await csv_state.clear()

                else:
                    param_temp[param]=str(form_data[param])
            else:
                result = {}
                # 遍历原始字典
                for key, value in form_data.items():
                    if key.startswith(param + '_'):
                        remaining_part = key[len(param)+1:]
                        parts = remaining_part.split('_')
                        
                        if len(parts) >= 2:
                            try:
                                num = int(parts[-1])
                            except ValueError:
                                continue 
                            new_key = '_'.join(parts[:-1])
                            
                            if num not in result:
                                result[num] = {}

                            if param_source["params"][param]['nested'][new_key]['type']=="bool":
                                if value=='False':
                                    value=False
                                elif value=='True':
                                    value=True
                                elif value=='on':
                                    value=True
                                else:
                                    value=str(value) 
                            else:
                                value=str(value)

                            result[num][new_key] = value
                
                for i in result:
                    for sub_param in param_source["params"][param]['nested']:
                        if sub_param not in result[i]:
                            result[num][sub_param] = False
                sorted_complex = [result[num] for num in sorted(result.keys())]
                param_temp[param]=sorted_complex
        
        if param_temp!=self.timeline[index]:
            self.timeline_edit(param_temp,sorted_complex,index)
            await self.refresh_iframe()
        point=int(form_data['add_or_move_point'].replace('index_',''))
        if "move_type" in form_data:
            move_type=form_data['move_type']
        else:
            move_type='Move'
        if point==-2:
            pass
        else:
            if move_type=='Move':
                self.timeline_delete(index,True)
                if point==-1:
                    self.timeline_push(param_temp, sorted_complex)
                else:
                    if point<=index:
                        self.timeline_push(param_temp, sorted_complex,point)
                    else:
                        point=point-1
                        self.timeline_push(param_temp, sorted_complex,point)
                    #before_name=form_data['add_or_move_point'].replace('Before the ',"")
                    #for i in range(len(self.timeline)):
                        #if self.timeline[i]['name']==before_name:
                            #self.timeline_push(param_temp, sorted_complex,i)
                            #break
                if plugin_name=='procedure-start':
                    for i in range(len(self.timeline)):
                        if self.timeline[i]['type']=='procedure-end' and self.timeline[i]['connect']==form_data['name']:
                            self.Unterminated_procedure_list=[x for x in self.Unterminated_procedure_list if x != self.timeline[i]['connect']] 
                            break
                await self.refresh_iframe()
            else:
                param_temp_t=copy.deepcopy(param_temp)
                sorted_complex_t=copy.deepcopy(sorted_complex)
                num=1
                while True:
                    if param_temp['name']+f"_c{num}" not in self.name_used:
                        param_temp_t['name']=param_temp['name']+f"_c{num}"
                        break
                    num+=1
                    
                
                if point==-1:
                    self.timeline_push(param_temp_t, sorted_complex_t)
                else:
                    self.timeline_push(param_temp_t, sorted_complex_t,point)
                await self.refresh_iframe()
        old_name=self.timeline[index]['name']
        if param_temp['name']!=old_name:
            self.update_next_name(param_temp['type'])
        


    

    


    @rx.event
    async def handle_expload(
        self, files: list[rx.UploadFile]
    ):
        if len(files)>0:
            for item in self.user_floder.iterdir():
                if item.is_dir() and item.name != "dist":
                    shutil.rmtree(item)
                elif not item.is_dir():
                    item.unlink()
                
            file=files[0]
            upload_data = await file.read()
            outfile = self.user_floder / file.name
            with outfile.open("wb") as file_object:
                file_object.write(upload_data)
            if file.name.endswith(".zip"):
                directory = os.path.dirname(str(outfile))
                with zipfile.ZipFile(str(outfile), 'r') as zip_ref:
                    zip_ref.extractall(directory)
                outfile.unlink()
                for item in self.user_floder.iterdir():
                    if item.is_dir() and item.name != "dist":
                        shutil.rmtree(item)
                jsexp_file=[file for file in self.user_floder.iterdir() if file.name.endswith('.jsexp')]
                if len(jsexp_file)>0:
                    await self.load_exp_json(jsexp_file[0])
                else:
                    await self.load_exp_json()
                
            else:
                await self.load_exp_json(outfile)
            



@rx.memo 
def flow_card(flow: dict[str, Any], index: int) -> rx.Component:
    cases = [
        (
            State.simple_params_state[pl]["type"],
            plugin_overlap_edit(plugin_name=pl, simple_params_by_plugin=flow, index=index,complex_list=State.complex_list[index])
        )
        for pl in simple_params
    ]
    def name_tooltip_text(flow: dict[str, Any],color:str=None):
        if color:
            return rx.tooltip(rx.text(flow["name"].to(str),align="center",justify="center",color= color),content=flow["type"].to(str))
        return rx.tooltip(rx.text(flow["name"].to(str),align="center",justify="center"),content=flow["type"].to(str))
    
    

    return rx.box(rx.cond((flow['type']==State.simple_params_state['procedure-start']["type"])|(flow['type']==State.simple_params_state['procedure-end']["type"]),
        rx.dialog.root(
        rx.dialog.trigger(
            
        rx.button(
            rx.cond(flow['type']==State.simple_params_state['procedure-start']["type"],
                rx.box(rx.hstack(rx.icon('arrow-right-from-line',align="center",justify="center"),rx.cond(
                flow['disabled'],name_tooltip_text(flow,color='gray'),name_tooltip_text(flow)
            ),align="center",justify="center",align_items="center",)),
            rx.box(rx.hstack(rx.cond(
            flow['disabled'],name_tooltip_text(flow,color='gray'),name_tooltip_text(flow)
        ),rx.icon('arrow-left-from-line',align="center",justify="center")),align="center",justify="center",align_items="center",)),
        size="2",
        style={
            "flex": "0 0 auto",
            "background_color": "light_blue",  
            "border": "1px solid light_blue",  
            "color": "white",  
            "border_radius": "8px",       
            "padding": "20px 20px",       
            "font_size": "18px",  
            "width": "fit-content",      
            "box_shadow": "0 1px 2px rgba(0,0,0,0.1)", 
            "margin_top": "12px",   
        }
    )
    
    ),
        rx.dialog.content(
                rx.match(flow.get("type"), *cases, rx.fragment()),key=f'dialog-edit-{index}'
                ,max_width="600px",
                max_height="650px",
                overflow_y="auto",
                force_mount=True ,on_close_auto_focus=Csv_state.clear,on_interact_outside=rx.prevent_default),force_mount=True),rx.dialog.root(
        rx.dialog.trigger(rx.button(
        rx.cond(
            flow['disabled'],name_tooltip_text(flow,color='gray'),name_tooltip_text(flow)
        ),
        size="2",
        style={
            "flex": "0 0 auto",
            "background_color": "white",  
            "border": "1px solid black",  
            "color": "black",  
            "border_radius": "8px",       
            "padding": "20px 20px",       
            "font_size": "18px",  
            "width": "fit-content",      
            "box_shadow": "0 1px 2px rgba(0,0,0,0.1)", 
            "margin_top": "12px",   
        }
    )),
        rx.dialog.content(
                rx.match(flow.get("type"), *cases, rx.fragment()),key=f'dialog-edit-{index}'
                ,max_width="600px",
                max_height="650px",
                overflow_y="auto",
                force_mount=True ,on_close_auto_focus=Csv_state.clear,on_interact_outside=rx.prevent_default),force_mount=True)))



def complex_form(complex,complex_index,plugin_name,name,complex_list,index=None):
    complex_inputs=[]
    param=params_dict[plugin_name]["params"][name]
    for param_sub in param['nested']:
        name_id=f'{name}_{param_sub}_{complex_index}'
        complex_inputs.extend(param_to_input(plugin_name,name, complex_list=complex_list,sub_name=param_sub,default_val=complex[param_sub],name_id=name_id,index=index))
        
    return rx.card(rx.vstack(*complex_inputs,rx.button(
                            "Delete",
                            variant="soft",
                            color_scheme="red",
                            on_click=lambda :complex_temp_state.delete_complex_temp(complex_index,index),type='button',width='100%'
                        ),width='100%'),width='100%',border="1px solid #ccc",)
    
def param_to_input(plugin_name,name,complex_list,sub_name=None,default_val='',name_id=None,index=None):                 
        if not sub_name:
            param=params_dict[plugin_name]["params"][name]
        else:
            param=params_dict[plugin_name]["params"][name]['nested'][sub_name]
            name=sub_name
        if not name_id:
            name_id=name
        label=rx.text(
                    f"{name.replace("_"," ").title()} ({'array of ' if param['array'] else ''}{param['type']})",
                    size="2",
                    margin_bottom="4px",
                    weight="bold",
                )

        
        if param["type"]!="complex":
            if name=="code":
                label=rx.text(
                    f"{name.replace("_"," ").title()} (code)",
                    size="2",
                    margin_bottom="4px",
                    weight="bold",
                )
            if name=="timeline_variables":
                label=rx.text(
                    f"{name.replace("_"," ").title()} (code)",
                    size="2",
                    margin_bottom="4px",
                    weight="bold",
                )
                input=rx.tabs.root(
                    rx.tabs.list(
                        rx.tabs.trigger("From code", value="tab1"),
                        rx.tabs.trigger("From csv/excel", value="tab2"),
                    ),
                    rx.tabs.content(
                        
                        rx.text_area(
                                name=name_id,
                                default_value=default_val,
                                width='100%',margin_bottom="4px",rows='10'
                            ),rx.callout(
                            rx.html("Use <em><b>jsPsych.timelineVariable(variable_name)</b></em> or <em><b>jsPsych.evaluateTimelineVariable(variable_name)</b></em> within the procedure."),
                            icon="info", size="1"
                        ),value="tab1"
                    ),
                    rx.tabs.content(
                        rx.upload(
                        rx.vstack(
                            rx.button(
                                "Select the csv/excel File",
                                color=color,
                                bg="white",
                                border=f"1px solid {color}",type='button'
                            ),
                            rx.text(
                                "Drag and drop .csv/.xlsx here or click to select the csv/excel file"
                            ),
                            
                           
                            rx.input(value=Csv_state.current_csv,name=name_id)
                            
                
                        ,align='center'
                            
                        ),
                        id="upload3",
                        border=f"1px dotted {color}",
                        padding="5em",
                        max_files=1,
                        accept={
                "application/csv": [".csv"],"application/xlsx": [".xlsx"]
            },no_keyboard=True,on_drop=Csv_state.handle_csv_upload(rx.upload_files(upload_id="upload3"))),
                        value="tab2",name=name_id+"_csv",
                    ),default_value="tab1",
                )
            elif name=='data_value':
                label=rx.text(
                    f"{name.replace("_"," ").title()} (code)",
                    size="2",
                    margin_bottom="4px",
                    weight="bold",
                )
                input=rx.tabs.root(
                    rx.tabs.list(
                        rx.tabs.trigger("From code", value="tab1"),
                        rx.tabs.trigger("From csv/excel", value="tab2"),
                    ),
                    rx.tabs.content(
                        
                        rx.text_area(
                                name=name_id,
                                default_value=default_val,
                                width='100%',margin_bottom="4px",rows='10'
                            ),value="tab1"
                    ),
                    rx.tabs.content(
                        rx.upload(
                        rx.vstack(
                            rx.button(
                                "Select the csv/excel File",
                                color=color,
                                bg="white",
                                border=f"1px solid {color}",type='button'
                            ),
                            rx.text(
                                "Drag and drop .csv/.xlsx here or click to select the csv/excel file"
                            ),
                            
                           
                            rx.input(value=Csv_state.current_csv,name=name_id)
                            
                
                        ,align='center'
                            
                        ),
                        id="upload3",
                        border=f"1px dotted {color}",
                        padding="5em",
                        max_files=1,
                        accept={
                "application/csv": [".csv"],"application/xlsx": [".xlsx"]
            },no_keyboard=True,on_drop=Csv_state.handle_csv_upload(rx.upload_files(upload_id="upload3"))),
                        value="tab2",name=name_id+"_csv",
                    ),default_value="tab1",
                )
            elif param["type"]=="bool":
                    input=rx.cond(
                        default_val,
                        rx.switch(
                                        name=name_id,
                                        default_checked=True,
                                        width='100%'
                                        
                                    ),
                        rx.switch(
                                        name=name_id,
                                        default_checked=False,
                                        width='100%'
                                        
                                    )

                    )
            elif param["type"]=="select":
                if name=='connect':
                    if str(index)=="None":
                        input=rx.select(
                                    State.Unterminated_procedure_list,
                                    name=name_id,width='100%',required=True
                                )
                    else:
                        input=rx.select.root(
                            rx.select.trigger(width='100%'),
                            rx.select.content(
                                    rx.select.item(default_val, value=default_val),
                                    rx.foreach(State.Unterminated_procedure_list,lambda val:rx.select.item(val, value=val)),width='100%'
                            ),
                            default_value=default_val,name=name_id,width='100%'
                        )
                else:
                    input=rx.select(
                                param["options"],
                                default_value=default_val,
                                name=name_id,width='100%'
                            )
            elif param["type"] in ("function",'object','html_string'): 
                input=rx.text_area(
                                name=name_id,
                                default_value=default_val,
                                    width='100%',rows='10'
                                
                                
                            )
            else:
                input=rx.cond(
                    default_val==State.false,
                    rx.input(
                                name=name_id,
                                default_value='False',
                                    width='100%'
                                
                            ),
                    rx.input(
                                name=name_id,
                                default_value=default_val,
                                    width='100%'
                                
                            )

                )
                    
            return [label,input]


        else:
            complex_form_card=rx.vstack(rx.foreach(
                complex_list,
                lambda complex, i: complex_form(complex=complex, complex_index=i,plugin_name=plugin_name,name=name,complex_list=complex_list,index=index)
            ),width='100%')
            
            add_complex=icon_button('plus',on_click= lambda :complex_temp_state.add_complex_temp(plugin_name,name,index))
                       
            return [label,complex_form_card,add_complex]


def plugin_form(plugin_name,simple_params_by_plugin,complex_list,index=None):
        
      
        plugin_params=params_dict[plugin_name]["params"]
        plugin_param_inputs=[]
        common_param_inputs=[]
        for param in plugin_params:
            #print(plugin_name,param)
            if "common" in plugin_params[param]:
                common_param_inputs.extend(param_to_input(plugin_name,param,default_val=simple_params_by_plugin[param],complex_list=complex_list,index=index))
            else:
                plugin_param_inputs.extend(param_to_input(plugin_name,param,default_val=simple_params_by_plugin[param],complex_list=complex_list,index=index))
        add_to=[]
        if str(index)=="None":
            heading=rx.heading("Plugin")
            add_to=[
                rx.text("Add Point",margin_bottom="4px",weight="bold"),
                rx.select.root(
                            rx.select.trigger(width='100%'),
                            rx.select.content(
                                    rx.foreach(State.timeline,lambda val,i:rx.select.item(f'Before the {val['name']}', value=f"index_{i}")),rx.select.item('At the end', value='index_-1'),width='100%'
                            ),
                            default_value='index_-1',name='add_or_move_point',width='100%'
                        )
            ]
        else:
            heading=rx.hstack(
                    rx.heading("Plugin"),
                    rx.hstack(
                        rx.text("Disabled"),
                        rx.checkbox(
                        name="disabled",
                        default_checked=simple_params_by_plugin['disabled'],
                    )),
                    justify="between",
                    width="100%"
                )
            if plugin_name in ['procedure-end','procedure-start']:
                add_to=[
                    rx.text("Move",margin_bottom="4px",weight="bold"),
                    rx.select.root(
                                rx.select.trigger(width='100%'),
                                rx.select.content(
                                        rx.select.item('Original', value='index_-2'),rx.foreach(State.timeline,lambda val,i:rx.select.item(f'Before the {val['name']}', value=f'index_{i}')),rx.select.item('At the end', value='index_-1'),width='100%'
                                ),
                                default_value='index_-2',name='add_or_move_point',width='100%'
                            )
                ]
    
            else:
                add_to=[
                    rx.text("Move",margin_bottom="4px",weight="bold"),
                    rx.select.root(
                                rx.select.trigger(width='100%'),
                                rx.select.content(
                                        rx.select.item('Original', value='index_-2'),rx.foreach(State.timeline,lambda val,i:rx.select.item(f'Before the {val['name']}', value=f'index_{i}')),rx.select.item('At the end', value='index_-1'),width='100%'
                                ),
                                default_value='index_-2',name='add_or_move_point',width='100%'
                            ),
                    rx.text("Move Type",margin_bottom="4px",weight="bold"),
                    rx.select.root(
                                rx.select.trigger(width='100%'),
                                rx.select.content(
                                        rx.select.item('Move', value='Move'),rx.select.item('Copy', value='Copy'),width='100%'
                                ),
                                default_value='Move',name='move_type',width='100%'
                            )
                ]
    
        if len(common_param_inputs)>0:
            section_common=[rx.section(
                                rx.heading("Common"),
                                *common_param_inputs,
                                padding_left="12px",
                                padding_right="12px",
                                background_color="var(--gray-2)",
                            )]
        else:
            section_common=[]
        plugin_dialog_box=rx.box(
                            rx.section(
                                heading,
                                rx.text("Name",margin_bottom="4px",weight="bold"),
                                rx.input(name="name",default_value=simple_params_by_plugin["name"]),
                                *add_to,
                                *plugin_param_inputs,
                                padding_left="12px",
                                padding_right="12px",
                                background_color="var(--gray-2)",
                            ),
                            *section_common,
                            width="100%",
                        )
                        
        return plugin_dialog_box

def plugin_overlap_add(plugin_name,simple_params_by_plugin,complex_list):
        if plugin_name in ['procedure-end','procedure-start']:
            description=rx.html(f"<a href='https://www.jspsych.org/v{version[0]}/overview/timeline/' target='_blank' rel='noopener noreferrer' style='color: #add8e6; text-decoration: underline;'>More information</a>")
        elif plugin_name in ['code']:
            description=rx.html(f"<a href='https://www.jspsych.org/v{version[0]}/reference/jspsych-data/' target='_blank' rel='noopener noreferrer' style='color: #add8e6; text-decoration: underline;'>More information</a>")
        elif plugin_name in ['data-variable']:
            description=rx.html(f"<a href='https://www.jspsych.org/v{version[0]}/reference/jspsych-randomization/' target='_blank' rel='noopener noreferrer' style='color: #add8e6; text-decoration: underline;'>More information</a>")
        else:
            description=rx.html(f"<a href='https://www.jspsych.org/v{version[0]}/plugins/{plugin_name}/' target='_blank' rel='noopener noreferrer' style='color: #add8e6; text-decoration: underline;'>More information</a>")
        plugin_dialog=rx.dialog.root(
            rx.dialog.trigger(
                rx.button(
                    plugin_name.replace("-", " ").title(),
                    width="90%",
                    height="auto",  
                    align_items="center",  
                    justify_content="center",
                    
                   
                )
            ),
            rx.dialog.content(
                rx.form(
                rx.hstack(
                    rx.dialog.title(f"{plugin_name.replace('-', ' ').title()}"),
                    rx.hstack(
                        rx.dialog.close(
                                    rx.button(
                                        "Cancel",
                                        variant="soft",
                                        color_scheme="gray",type='button'
                                    ),
                                ),
                        rx.dialog.close(
                                    rx.button(
                                        "Add",
                                        type="submit",
                                    )
                                )),
                    justify="between",
                    width="100%"
                ),
                rx.dialog.description(''),
                description,
                    rx.flex(plugin_form(plugin_name,simple_params_by_plugin,complex_list),
                        rx.flex(
                            rx.dialog.close(
                                rx.button(
                                    "Cancel",
                                    variant="soft",
                                    color_scheme="gray",type='button'
                                ),
                            ),
                            rx.dialog.close(
                                rx.button(
                                    "Add",
                                    type="submit",
                                )
                            ),
                            spacing="3",
                            justify="end",
                        ),
                        direction="column",
                        spacing="4",
                    ),
                    on_submit=lambda data: State.add_plugin(plugin_name, data),
                    id="outer-form",
                    reset_on_submit=False,
                ),
                max_width="600px",
                max_height="650px",
                overflow_y="auto",on_close_auto_focus=[complex_temp_state.clear_complex_temp,Csv_state.clear],on_interact_outside=rx.prevent_default
            )
        )
        return plugin_dialog


def plugin_overlap_edit(plugin_name: str,simple_params_by_plugin: dict[str, Any],index: int,complex_list:list[dict[str, Any]]):
        if plugin_name in ['procedure-end','procedure-start']:
            description=rx.html(f"<a href='https://www.jspsych.org/v{version[0]}/overview/timeline/' target='_blank' rel='noopener noreferrer' style='color: #add8e6; text-decoration: underline;'>More information</a>")
        elif plugin_name in ['code']:
            description=rx.html(f"<a href='https://www.jspsych.org/v{version[0]}/reference/jspsych-data/' target='_blank' rel='noopener noreferrer' style='color: #add8e6; text-decoration: underline;'>More information</a>")
        elif plugin_name in ['data-variable']:
            description=rx.html(f"<a href='https://www.jspsych.org/v{version[0]}/reference/jspsych-randomization/' target='_blank' rel='noopener noreferrer' style='color: #add8e6; text-decoration: underline;'>More information</a>")
        else:
            description=rx.html(f"<a href='https://www.jspsych.org/v{version[0]}/plugins/{plugin_name}/' target='_blank' rel='noopener noreferrer' style='color: #add8e6; text-decoration: underline;'>More information</a>")
        
        plugin_dialog=rx.box(rx.form(rx.hstack(
                    rx.dialog.title(f"{plugin_name.replace('-', ' ').title()}"),
                    rx.hstack(
                        rx.dialog.close(
                                rx.button(
                                    "Cancel",
                                    variant="soft",
                                    color_scheme="gray",type='button'
                                ),
                            ),rx.dialog.close(
                        rx.button(
                            "Delete",
                            variant="soft",
                            color_scheme="red",
                            on_click=lambda :State.delete_plugin(index),type='button'
                        ),
                    ),rx.dialog.close(
                                rx.button(
                                    "Edit",
                                    type="submit"
                                )
                            )),
                    justify="between",
                    width="100%"
                ),
                rx.dialog.description(''),
                description,                
                    rx.flex(plugin_form(plugin_name,simple_params_by_plugin,complex_list,index=index),
                        rx.flex(
                            rx.dialog.close(
                                rx.button(
                                    "Cancel",
                                    variant="soft",
                                    color_scheme="gray",type='button'
                                ),
                            ),
                            
                            rx.dialog.close(
                                rx.button(
                                    "Edit",
                                    type="submit"
                                )
                            ),
                            spacing="3",
                            justify="end",
                        ),
                        direction="column",
                        spacing="4",
                    ),
                    on_submit=lambda data: State.edit_plugin(plugin_name, data,index),
                    reset_on_submit=False,
                ))
                
            
        return plugin_dialog

@rx.memo
def plugin_button_list() -> rx.Component:
    plugin_button_list=[]
    special_button_list=[]
    for plugin_name in params_dict:
        plugin_dialog=plugin_overlap_add(plugin_name,State.simple_params_state[plugin_name],complex_temp_state.complex_list_temp)
        if plugin_name in ['procedure-end','procedure-start']:
            special_button_list.append(plugin_dialog)
        else:
            plugin_button_list.append(plugin_dialog)
    
    return rx.box(
                rx.heading("Procedure", margin_bottom="1rem"),
                rx.vstack(
                    *special_button_list,
                    align="center",
                    height="10%",
                    overflow_y="auto",
                    spacing="2",  
                ),
                rx.heading("Plugin", margin_bottom="1rem"),
                rx.vstack(
                    *plugin_button_list,
                    align="center",
                    height="75%",
                    overflow_y="auto",
                    spacing="2",  
                ),
                align="center",
                
                padding="1rem",
                flex_basis="20%",
                border="1px solid #ccc",
                border_radius="8px",
                box_shadow="0 2px 6px rgba(0,0,0,0.1)",
                min_width="150px",
                height="100%",
                spacing="1"
            )

@rx.memo
def timeline_flow(timeline:list[dict[str, Any]]) -> rx.Component:
    return rx.box(
            rx.heading("Timeline"),
            rx.hstack(
                rx.foreach(
                timeline,
                lambda flow, i: flow_card(flow=flow, index=i)
                ),
                spacing="1",
                overflow_x="auto",
                flex_wrap="nowrap",
                max_width="100%",
                style={
                    "scrollbar-width": "thin",  # Firefox
                }
             ),
            height="17vh",
            padding="1rem",
            border="1px solid #ccc",
            border_radius="8px",
            box_shadow="0 2px 6px rgba(0,0,0,0.1)",
            width="100%",
            overflow_y="auto",
        )

@rx.memo
def preview_win(iframe_key:int)  -> rx.Component:
    return rx.box(
                    rx.hstack(rx.hstack(
                        rx.heading("Preview"),
                        rx.button(rx.box(style=circle_style), on_click=State.refresh_iframe_simple,
                                    padding="0",             
                                    border="none",          
                                    bg="transparent",     
                                    width="32px",           
                                    height="32px",          
                                    _hover={"bg": "transparent"}, 
                                    _active={"bg": "transparent"})),rx.hstack(rx.select.root(
                            rx.select.trigger(width='100%'),
                            rx.select.content(
                                    rx.foreach(State.timeline,lambda val,i:rx.select.item(f'{val['name']}', value=f"index_{i}",width='100%')),width='100%',
                            ),
                            default_value='index_0',value=State.preview_start,name='preview_start',width='20%',on_change=[State.set_preview_start,State.refresh_iframe],
                        )), justify="between",width='100%'),
                rx.el.iframe(
                    src=f"{rx.get_upload_url(State.preview_file)}?id={iframe_key}",
                    width="100%",
                    height="90%",
                    loading='eager',allow='fullscreen',background_color="white" 
                ),
                flex_basis="90%",
                border="1px solid #ccc",
                border_radius="8px",
                box_shadow="0 2px 6px rgba(0,0,0,0.1)",
                height="90%",width="100%",padding="1rem",force_mount=True)

def icon_button(icon:str,on_click,type='button') -> rx.Component:
    return rx.button(
            rx.icon(icon),
            style={
                'color':'black',
                "border": "1px solid #ccc",      
                "height": "60%",           
                "width": "auto",              
                "aspect_ratio": "1 / 1",      
                "display": "flex",
                "align_items": "center",
                "justify_content": "center",
                "cursor": "pointer",
                "box_shadow": "0 1px 2px rgba(0,0,0,0.2)",
                'align':"center"
            },color_scheme='gray',variant='outline',radius='full',on_click=on_click,type=type
        )

def icon_text_button(icon:str,text,on_click=None) -> rx.Component:
    return rx.button(
            rx.icon(icon),rx.text(text),
            style={
                'color':'black',
                "border": "1px solid #ccc",
                "height": "80%",           
                "width": "15%",                  
                "display": "flex",
                "align_items": "center",
                "justify_content": "center",
                "cursor": "pointer",
                "box_shadow": "0 1px 2px rgba(0,0,0,0.2)",
            },color_scheme='gray',variant='outline',radius='full',on_click=on_click
        )

@rx.memo
def setting_form(exp_name:str,plugin_source:str,data_save:str,head_script:str) -> rx.Component:
    return rx.dialog.root(
            rx.dialog.trigger(
                icon_text_button(icon='settings',text='Experiment Setting')
            ),
            rx.dialog.content(
                rx.hstack(
                    rx.dialog.title("Experiment Settings"),
                    rx.hstack(
                        rx.dialog.close(
                                    rx.button(
                                        "Cancel",
                                        variant="soft",
                                        color_scheme="gray",type='button'
                                    ),
                                ),
                        ),
                    justify="between",
                    width="100%"
                ),
                
                rx.dialog.description("Fill in to set the experiment settings"),
                rx.form(
                    rx.flex(
                        rx.text("Experiment Name", 
                        size="2",
                        margin_top="4px",
                        weight="bold",),
                        rx.input(
                            default_value=exp_name,
                            name='exp_name'
                        ),
                        rx.text("Plugin Source", 
                        size="2",
                        margin_top="4px",
                        weight="bold",),
                        rx.select(
                            ["Web", "Local", "NAODAO", "Credamo"],
                            default_value=plugin_source,
                            name='plugin_source'
                        ),
                        rx.text("Data Save", 
                        size="2",
                        margin_top="4px",
                        weight="bold",),
                        rx.select(
                            ["Display", "Local", "NAODAO", "Credamo",'JATOS','Local_Server'],
                            default_value=data_save,
                            name='data_save'
                        ),
                        rx.text("Head Script", 
                        size="2",
                        margin_top="4px",
                        weight="bold",),
                        rx.text_area(
                            default_value=head_script
                            ,name='head_script',rows='10'
                        ),
                        rx.flex(
                            rx.dialog.close(
                                rx.button("Save", color_scheme="blue",type='submit'),
                            ),
                            justify="end",
                            spacing="3"
                        ), 
                        spacing="2",
                        direction="column",
                    ),
                
                on_submit=State.save_exp_settings,
                reset_on_submit=True,
                ),
                max_width="450px",on_interact_outside=rx.prevent_default
            ),
        
        )
color = "rgb(107,99,246)"

@rx.memo
def new_exp_form() -> rx.Component:
    return rx.dialog.root(
            rx.dialog.trigger(
                icon_text_button(icon='file-plus-2',text='New Experiment')
            ),
            rx.dialog.content(
                 rx.hstack(
                    rx.dialog.title("New Experiment"),
                    rx.hstack(
                        rx.dialog.close(
                                    rx.button(
                                        "Cancel",
                                        variant="soft",
                                        color_scheme="gray",type='button'
                                    ),
                                ),
                        ),
                    justify="between",
                    width="100%"
                ),
                
                rx.dialog.description("Click the button to create a new experiment"),
                rx.callout(
                            "This will clean the current experiment!",
                            icon="info", size="1",color_scheme='red',width="100%",margin_top='10px',margin_bottom='10px'
                        ),
                rx.flex(
                            rx.dialog.close(rx.button(
                            "New Experiment",
                            on_click=[State.new_exp,Uploaded_files_state.delete_all_file],margin_top='10px'
                                    )), 
                            
                            justify="start",
                            spacing="3"
                        )
                
                    ,max_width="450px"))

                    
                

@rx.memo
def load_exp_form() -> rx.Component:
    return rx.dialog.root(
            rx.dialog.trigger(
                icon_text_button(icon='cloud-upload',text='Load Experiment')
            ),
            rx.dialog.content(
                
                rx.hstack(
                    rx.dialog.title("Load Experiment"),
                    rx.hstack(
                        rx.dialog.close(
                                    rx.button(
                                        "Cancel",
                                        variant="soft",
                                        color_scheme="gray",type='button'
                                    ),
                                ),
                        ),
                    justify="between",
                    width="100%"
                ),
                rx.dialog.description("Select .zip or .jsexp to load experiment"),
                rx.callout(
                            "This will overwrite the current experiment!",
                            icon="info", size="1",color_scheme='red',width="100%",margin_top='5px'
                        ),
                rx.vstack(
                    rx.upload(
                        rx.vstack(
                            rx.button(
                                "Select the Experiment File",
                                color=color,
                                bg="white",
                                border=f"1px solid {color}",
                            ),
                            rx.text(
                                "Drag and drop .zip or .jsexp here or click to select the experiment file"
                            ),
                            
                            rx.foreach(
                                rx.selected_files("upload2"), rx.card
                            )
                
                        ,align='center'
                            
                        ),
                        id="upload2",
                        border=f"1px dotted {color}",
                        padding="5em",
                        max_files=1,
                        accept={
                "application/zip": [".zip"],
                "application/json": [".jsexp"],
            },no_keyboard=True,
                    ),
                    rx.flex(
                        rx.dialog.close(rx.button(
                            "Load",
                            on_click=[
                                Uploaded_files_state.delete_all_file,
                                State.handle_expload(rx.upload_files(upload_id="upload2")),
                                rx.clear_selected_files("upload2")
                            ],
                        ))
                        ,justify="start",spacing="3",width='100%'),

                    
                    
                    width="100%",margin_top='5px'
                ),
                max_width="450px",
                max_height="550px"
            ),
        )

@rx.memo
def save_exp_form(exp_file:str) -> rx.Component:
    return rx.dialog.root(
            rx.dialog.trigger(
                icon_text_button(icon='cloud-download',text='Save Experiment'),on_click=State.save_exp_file
            ),
            rx.dialog.content(
                 rx.hstack(
                    rx.dialog.title("Save Experiment"),
                    rx.hstack(
                        rx.dialog.close(
                                    rx.button(
                                        "Cancel",
                                        variant="soft",
                                        color_scheme="gray",type='button'
                                    ),
                                ),
                        ),
                    justify="between",
                    width="100%"
                ),
                
                rx.dialog.description("Download your Experiment"),
                rx.flex(
                            rx.cond(
                                State.building_exp_file,
                                rx.button('Preparing',on_click=rx.download(url=rx.get_upload_url(exp_file)),target='_blank',disabled=True,color_scheme='gray',margin_top='10px'),
                                rx.dialog.close(
                                rx.button('Download',on_click=rx.download(url=rx.get_upload_url(exp_file)),target='_blank',disabled=False,margin_top='10px'))
                                
                                ),
                            justify="start",
                            spacing="3"
                        ), 
                
                
            
                max_width="450px",force_mount=True
            ),force_mount=True
        )

@rx.memo
def file_list_item(name:str,id:int):
    return rx.hstack(
        rx.card(rx.hstack(rx.hstack(
        rx.icon('file-check'),rx.text(name)),rx.button("Delete",on_click=lambda: Uploaded_files_state.delete_file(name,id),),justify='between',width='100%',align='center'),width='100%'),width='100%'
    )



@rx.memo
def load_files_form() -> rx.Component:
    return rx.dialog.root(
            rx.dialog.trigger(
                icon_text_button(icon='image-up',text='Load Material Files')
            ),
            rx.dialog.content(
                
                rx.hstack(
                    rx.dialog.title("Load Material Files"),
                    rx.hstack(
                        rx.dialog.close(
                                    rx.button(
                                        "Close",
                                        variant="soft",
                                        color_scheme="gray",type='button'
                                    )
                                ),
                        ),
                    justify="between",
                    width="100%"
                ),
                rx.dialog.description("Load material files used in experiment"),
                rx.vstack(
                    rx.upload(
                        rx.vstack(
                            rx.button(
                                "Select Files and Upload",
                                color=color,
                                bg="white",
                                border=f"1px solid {color}",
                            ),
                            rx.text(
                                "Drag and drop files here or click to select files and upload"
                            ),
                            
                            rx.foreach(
                                rx.selected_files("upload1"), rx.card
                            )
                
                        ,align='center'
                            
                        ),
                        id="upload1",
                        border=f"1px dotted {color}",
                        padding="5em",no_keyboard=True,on_drop=Uploaded_files_state.handle_upload(rx.upload_files(upload_id="upload1"))
                    ),
                    rx.flex(
                        #rx.button(
                            #"Upload",
                          #  on_click=[
                          #      State.handle_upload(rx.upload_files(upload_id="upload1")),
                          #      rx.clear_selected_files("upload1")
                          #  ],
                       # ),
                        rx.button(
                            "Delete All Files",
                            on_click=[Uploaded_files_state.delete_all_file,rx.clear_selected_files("upload1")],
                        ),justify="start",spacing="3",width='100%'),
                    
                    rx.foreach(
                                Uploaded_files_state.files,
                                lambda name,id: file_list_item(name=name,id=id),
                            ),
                    
                    
                    padding="5em",width="100%",
                ),
                max_width="600px",
                max_height="700px"
            ),
        )

@rx.memo
def commend_box() -> rx.Component:
    return rx.box(rx.hstack(
        setting_form(exp_name=State.exp_name,plugin_source=State.plugin_source,data_save=State.data_save,head_script=State.head_script),
        new_exp_form(),
        load_exp_form(),
        save_exp_form(exp_file=State.exp_file),
        load_files_form(),
        rx.link(rx.button(
            rx.icon('external-link'),rx.text('External Preview'),color_scheme='gray',variant='outline',radius='full',
            style={
                'color':'black',
                "border": "1px solid #ccc",
                "height": "100%",           
                "width": "100%",                  
                "display": "flex",
                "align_items": "center",
                "justify_content": "center",
                "cursor": "pointer",
                "box_shadow": "0 1px 2px rgba(0,0,0,0.2)",
            }
            ),height="80%",     
                width="15%",href=rx.get_upload_url(State.preview_file), underline='none',target='_blank'
            ),
        align="center",
        justify="center",
        align_items="center",
        height="100%",
        spacing='3'  ,overflow_y="auto"
        
    ),  border="1px solid #ccc",
        border_radius="8px",
        box_shadow="0 2px 6px rgba(0,0,0,0.1)",
        height="10%",
        width="100%",
        justify_content="flex-start",
        align_items="center",padding_left='1rem',padding_right='1rem')

@rx.page(on_load=State.set_user_floder,title='jsPsych Easy Builder' )
def index():
    return rx.vstack(
                rx.hstack(
                    plugin_button_list(),
                    rx.vstack(
                        preview_win(iframe_key=Preview_state.iframe_key),
                        commend_box(),
                        height="100%",
                        width="100%",
                        ), 
                    height="82vh",  
                    width="100%",)
                ,timeline_flow(timeline=State.timeline),
                height="100vh",
                width="100%",
                background_color="white" 
    )
    

if mode=='Local':
    app = rx.App(api_transformer=data_upload)
else:
    app = rx.App()
